import json
import io
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import MonthPlan, DailyRecord
from django.db.models.functions import Coalesce
from django.db.models import Sum


def index(request):
    plans = MonthPlan.objects.all()[:12]

    plans_json = list(
        MonthPlan.objects
        .annotate(
            fact_sales=Coalesce(Sum('daily_records__sales'), 0),
            fact_leads=Coalesce(Sum('daily_records__leads'), 0),
        )
        .values(
            'id',
            'month',
            'plan_sales',
            'plan_leads',
            'fact_sales',
            'fact_leads',
            'work_days',
            'passed_days',
        )[:12]
    )

    return render(request, 'index.html', {
        'plans': plans,
        'plans_json': plans_json,
    })


def dashboard(request, plan_id):
    plan = get_object_or_404(MonthPlan, id=plan_id)
    records = list(plan.daily_records.values(
        'day_number', 'date', 'sales', 'leads', 'deals', 'spend', 'comment'
    ))
    for r in records:
        if r['date']:
            r['date'] = str(r['date'])
    return render(request, 'dashboard.html', {
        'plan': plan,
        'records': plan.daily_records.all(),
        'records_json': json.dumps(records),
    })

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_dashboard(request, plan_id):
    try:
        plan = get_object_or_404(MonthPlan, id=plan_id)
        month = plan.month
        plan.delete()
        return JsonResponse({'status': 'ok', 'deleted_month': month})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
def save_dashboard(request):
    try:
        data = json.loads(request.body)
        month = data.get('month', '')
        plan, created = MonthPlan.objects.update_or_create(
            month=month,
            defaults={
                'plan_sales': int(data.get('plan_sales') or 0),
                'plan_leads': int(data.get('plan_leads') or 0),
                'work_days': int(data.get('work_days') or 26),
                'passed_days': int(data.get('passed_days') or 0),
            }
        )
        plan.daily_records.all().delete()
        for i, row in enumerate(data.get('rows', [])):
            date_val = row.get('date') or None
            DailyRecord.objects.create(
                month_plan=plan,
                day_number=i + 1,
                date=date_val,
                sales=int(row.get('sales') or 0),
                leads=int(row.get('leads') or 0),
                deals=int(row.get('deals') or 0),
                spend=int(row.get('spend') or 0),
                comment=row.get('comment', ''),
            )
        return JsonResponse({'status': 'ok', 'plan_id': plan.id, 'created': created})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_http_methods(["GET"])
def load_dashboard(request, month):
    try:
        plan = MonthPlan.objects.get(month=month)
        records = list(plan.daily_records.values(
            'day_number', 'date', 'sales', 'leads', 'deals', 'spend', 'comment'
        ))
        for r in records:
            if r['date']:
                r['date'] = str(r['date'])
        return JsonResponse({
            'status': 'ok',
            'plan_sales': plan.plan_sales,
            'plan_leads': plan.plan_leads,
            'work_days': plan.work_days,
            'passed_days': plan.passed_days,
            'rows': records
        })
    except MonthPlan.DoesNotExist:
        return JsonResponse({'status': 'not_found'}, status=404)


def export_excel(request, plan_id):
    plan = get_object_or_404(MonthPlan, id=plan_id)
    records = plan.daily_records.all()

    wb = Workbook()

    hdr_fill = PatternFill('solid', fgColor='01696F')
    hdr_font = Font(bold=True, color='FFFFFF', size=12)
    bold = Font(bold=True, size=11)

    ws1 = wb.active
    ws1.title = 'Настройки'
    for col, title in enumerate(['Параметр', 'Значение'], 1):
        c = ws1.cell(row=1, column=col, value=title)
        c.font = hdr_font
        c.fill = hdr_fill

    settings = [
        ('Месяц', plan.month),
        ('План продаж (сум)', plan.plan_sales),
        ('План лидов', plan.plan_leads),
        ('Рабочих дней', plan.work_days),
        ('Прошло дней', plan.passed_days),
    ]
    for i, (k, v) in enumerate(settings, start=2):
        ws1.cell(row=i, column=1, value=k).font = bold
        ws1.cell(row=i, column=2, value=v)

    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet('Дневные данные')
    headers = ['День', 'Дата', 'Продажи (сум)', 'Лиды', 'Сделки', 'Расход (сум)', 'Комментарий']
    widths = [8, 14, 18, 10, 10, 18, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col)].width = w

    for r in records:
        ws2.append([r.day_number, str(r.date) if r.date else '', r.sales, r.leads, r.deals, r.spend, r.comment])

    ws3 = wb.create_sheet('Итоги')
    total_sales = sum(r.sales for r in records)
    total_leads = sum(r.leads for r in records)
    total_deals = sum(r.deals for r in records)
    total_spend = sum(r.spend for r in records)
    conversion = round(total_deals / total_leads * 100, 1) if total_leads else 0
    plan_pct = round(total_sales / plan.plan_sales * 100, 1) if plan.plan_sales else 0

    ws3['A1'] = 'Итоговый отчёт'
    ws3['A1'].font = Font(bold=True, size=14, color='01696F')

    summary = [
        ('Месяц', plan.month),
        ('Факт продаж', total_sales),
        ('План продаж', plan.plan_sales),
        ('Выполнение плана %', f'{plan_pct}%'),
        ('Факт лидов', total_leads),
        ('Сделки всего', total_deals),
        ('Конверсия %', f'{conversion}%'),
        ('Расход всего', total_spend),
    ]
    for i, (k, v) in enumerate(summary, start=3):
        ws3.cell(row=i, column=1, value=k).font = bold
        ws3.cell(row=i, column=2, value=v)

    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"dashboard_{plan.month}.xlsx\"'
    return response